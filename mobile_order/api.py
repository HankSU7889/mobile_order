import frappe
from typing import Optional, List, Dict, Any


@frappe.whitelist(allow_guest=True)
def get_item_index() -> Dict[str, Any]:
    if hasattr(frappe.local, 'mobile_order_item_index'):
        return frappe.local.mobile_order_item_index

    # 只获取变体产品（variant_of IS NOT NULL 且非空），排除模板
    items = frappe.get_all(
        "Item",
        filters={"disabled": 0, "variant_of": ["not like", ""]},
        fields=["item_code", "item_name", "item_group", "brand", "description"]
    )

    index = {}
    for item in items:
        if not item.item_name:
            continue
        parts = item.item_name.split("-")
        series = parts[0] if parts else item.item_name

        if series not in index:
            index[series] = []
        index[series].append({
            "item_code": item.item_code,
            "item_name": item.item_name,
            "item_group": item.item_group or "",
            "brand": item.brand or "",
            "description": item.description or ""
        })

    frappe.local.mobile_order_item_index = index
    return index


@frappe.whitelist(allow_guest=True)
def search_items(keyword: str) -> List[Dict[str, Any]]:
    if not keyword or len(keyword) < 1:
        return []

    keyword = keyword.strip()

    # 搜索变体产品（排除模板），同时匹配 item_name 和 description
    items = frappe.db.sql("""
        SELECT item_code, item_name, item_group, brand, description
        FROM tabItem
        WHERE disabled = 0
        AND variant_of IS NOT NULL AND variant_of != ''
        AND (item_name LIKE %s OR description LIKE %s)
        ORDER BY item_name
    """, ("%" + keyword + "%", "%" + keyword + "%"), as_dict=1)

    return items


@frappe.whitelist(allow_guest=True)
def get_all_series() -> List[str]:
    index = get_item_index()
    return sorted(index.keys())


@frappe.whitelist(allow_guest=True)
def get_items_by_series(series: str) -> List[Dict[str, Any]]:
    if not series:
        return []
    index = get_item_index()
    return index.get(series, [])


@frappe.whitelist(allow_guest=True)
def get_item_by_code(item_code: str) -> Optional[Dict[str, Any]]:
    if not item_code:
        return None
    try:
        item = frappe.get_doc("Item", item_code)
        return {
            "item_code": item.item_code,
            "item_name": item.item_name,
            "item_group": item.item_group,
            "brand": item.brand,
            "description": item.description or ""
        }
    except Exception:
        return None


@frappe.whitelist(allow_guest=True)
def get_sales_persons(active_only: bool = True) -> List[Dict[str, Any]]:
    filters = {}
    if active_only:
        filters["enabled"] = 1
    sales_persons = frappe.get_all(
        "Sales Person",
        filters=filters,
        fields=["name", "sales_person_name", "employee", "department", "enabled"],
        order_by="sales_person_name"
    )
    return sales_persons


@frappe.whitelist(allow_guest=True)
def create_customer_order(data: Dict[str, Any]) -> Dict[str, Any]:
    try:
        so = frappe.get_doc({
            "doctype": "Sales Order",
            "customer_name": data.get("customer_name"),
            "naming_series": "SO-MO-",
            "order_type": "Sales",
            "items": []
        })

        for item_data in data.get("items", []):
            so.append("items", {
                "item_code": item_data.get("item_code"),
                "qty": item_data.get("qty", 1),
                "rate": item_data.get("rate", 0),
            })

        so.insert()
        so.submit()

        return {
            "success": True,
            "order_id": so.name,
            "message": "订单 {0} 已创建".format(so.name)
        }
    except Exception as e:
        return {
            "success": False,
            "order_id": None,
            "message": str(e)
        }


@frappe.whitelist(allow_guest=True)
def get_customer_orders(sales_person: str = None, status: str = None) -> List[Dict[str, Any]]:
    filters = {}
    if sales_person:
        filters["sales_person"] = sales_person
    if status:
        filters["status"] = status

    orders = frappe.get_all(
        "Sales Order",
        filters=filters,
        fields=["name", "customer_name", "mobile_no", "sales_person", "status", "grand_total", "creation"],
        order_by="creation desc",
        limit=100
    )
    return orders


# =====================================================================
# 产品系列名称列表（来源于 物品导出2026年04月15日）
# =====================================================================
PRODUCT_SERIES = [
    "晶品系列", "雅派系列", "XFlip折叠", "晶钢系列", "格耀磁吸折叠", "晶耀支架", "御盾系列", "星尚系列", "格耀系列",
    "护眼系列", "商务系列", "幻彩系列", "炫透系列", "乐享短绳", "尊享系列", "钛色系列", "奢派系列", "本色系列",
    "炫透折叠", "晶砂磁吸", "秒变系列", "臻品系列", "酷纱系列", "至尊折叠", "晶耀苹果", "支点系列", "手腕挂绳",
    "潮变系列", "潮酷折叠", "奢盾系列", "晶闪系列", "星辰系列", "匠造磁吸", "至臻折叠", "星品系列", "钛金系列",
    "晶砂系列", "左右折叠", "强磁系列", "晶耀系列", "艺术系列", "非凡系列", "潮酷系列", "格耀支架折叠",
    "典范系列", "水晶之恋", "格耀磁吸", "晶透系列", "水晶系列", "奢尚系列", "奢品系列", "乐享长绳", "至臻系列",
    "睿享系列", "君尚系列", "晶砂支架", "名仕系列", "臻品折叠", "支盾系列", "晶盾系列", "强磁-", "炫变系列",
    "小香系列", "光晖支架", "晶耀磁吸苹果", "爱心挂绳", "舒感华为", "名匠折叠", "酷盾系列", "晶闪磁吸", "星钻系列",
    "匠造支架", "酷纱华为", "水晶折叠", "名匠系列", "上下折叠", "晶甲系列", "格耀基础折叠", "晶耀磁吸", "智享系列",
    "原款系列", "防窥钢化膜", "宝盒系列", "光晖折叠", "绅士系列", "格耀360支架", "Airtag防丢器", "镀纱系列", "典雅折叠",
    "名媛系列", "悦享系列", "晶砂折叠", "臻享系列", "明睿系列", "秒变XT系列", "舒感系列", "冰透系列",
    "御享系列", "臻皮系列", "光晖基础", "君睿系列", "晶耀360支架", "高定系列", "匠造基础", "至尊系列", "光晖磁吸",
    "水晶磁吸", "高奢系列", "典雅系列", "绅士折叠"
]


@frappe.whitelist(allow_guest=True)
def get_inventory_items(
    keyword: str = "",
    product_series: str = "",
    brand: str = "",
    status: str = "",   # "active" | "discontinued" | ""
    page: int = 1,
    page_size: int = 200
) -> Dict[str, Any]:
    """
    获取库存物料列表，支持：
    - keyword: 搜索物料名称/编码/品牌
    - product_series: 按产品系列筛选（对应 item_group）
    - brand: 按品牌筛选
    - status: "active"=在售(disabled=0), "discontinued"=停售(disabled=1), ""=全部
    """
    filters = []
    params = []

    if status == "active":
        filters.append("disabled = %s")
        params.append(0)
    elif status == "discontinued":
        filters.append("disabled = %s")
        params.append(1)
    # status="" 时默认只看启用（不显示禁用的）

    if keyword:
        filters.append("(item_name LIKE %s OR item_code LIKE %s OR brand LIKE %s)")
        params.extend(["%" + keyword + "%", "%" + keyword + "%", "%" + keyword + "%"])

    if product_series:
        filters.append("item_group = %s")
        params.append(product_series)

    if brand:
        filters.append("brand = %s")
        params.append(brand)

    where_clause = " AND ".join(filters) if filters else "disabled = 0"

    total = frappe.db.sql(
        f"SELECT COUNT(*) as cnt FROM tabItem WHERE {where_clause}",
        tuple(params), as_dict=1
    )[0].cnt

    offset = (page - 1) * page_size
    items = frappe.db.sql(
        f"""
        SELECT item_code, item_name, item_group, brand, description, disabled, modified
        FROM tabItem
        WHERE {where_clause}
        ORDER BY modified DESC
        LIMIT %s OFFSET %s
        """,
        tuple(params + [page_size, offset]), as_dict=1
    )

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@frappe.whitelist(allow_guest=True)
def get_inventory_item_groups() -> List[str]:
    """返回产品系列名称列表（固定来源于物品导出表格）"""
    return sorted(PRODUCT_SERIES)


@frappe.whitelist(allow_guest=True)
def get_inventory_brands() -> List[str]:
    """
    从 item_name 解析品牌：
    - 格式：{series_prefix}{brand}-{model}-{color}，如"光晖基础三星-S26-橙"
    - 规则：在第一段中找 PRODUCT_SERIES 最长前缀匹配，剩余部分=品牌
    - 精确匹配系列前缀时（如"光晖折叠"），品牌为空
    """
    items = frappe.db.sql("""
        SELECT DISTINCT item_name
        FROM tabItem
        WHERE disabled = 0
        AND item_name IS NOT NULL AND item_name != ""
        AND item_name LIKE "%-%-%"
        LIMIT 5000
    """)

    brands = set()
    for (item_name,) in items:
        parts = item_name.split("-")
        if len(parts) < 2:
            continue
        first = parts[0]
        # 找最长匹配的前缀
        best = None
        for series in PRODUCT_SERIES:
            if first.startswith(series):
                if best is None or len(series) > len(best):
                    best = series
        if best:
            brand = first[len(best):]
            if brand:
                brands.add(brand)
    return sorted(brands)


@frappe.whitelist(allow_guest=True)
def export_inventory_excel(
    keyword: str = "",
    product_series: str = "",
    brand: str = "",
    status: str = ""
) -> Dict[str, Any]:
    """导出库存数据为 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from frappe.utils import now_datetime
    import io

    filters = []
    params = []

    if status == "active":
        filters.append("disabled = %s")
        params.append(0)
    elif status == "discontinued":
        filters.append("disabled = %s")
        params.append(1)

    if keyword:
        filters.append("(item_name LIKE %s OR item_code LIKE %s OR brand LIKE %s)")
        params.extend(["%" + keyword + "%", "%" + keyword + "%", "%" + keyword + "%"])

    if product_series:
        filters.append("item_group = %s")
        params.append(product_series)

    if brand:
        filters.append("brand = %s")
        params.append(brand)

    where_clause = " AND ".join(filters) if filters else "disabled = 0"

    items = frappe.db.sql(
        f"""
        SELECT item_code, item_name, item_group, brand, description, disabled, modified
        FROM tabItem
        WHERE {where_clause}
        ORDER BY modified DESC
        """,
        tuple(params), as_dict=1
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存数据"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1890FF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["物料编码", "物料名称", "产品系列", "品牌", "描述", "状态", "最后修改"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in items:
        ws.append([
            row.item_code,
            row.item_name,
            row.item_group or "",
            row.brand or "",
            row.description or "",
            "停售" if row.disabled else "在售",
            row.modified.strftime("%Y-%m-%d %H:%M") if row.modified else ""
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventory_export_{now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    from frappe.utils.file_manager import save_file
    file = save_file(filename, output.getvalue(), "File", None, is_private=0)

    return {"file_url": file.file_url, "filename": filename}
